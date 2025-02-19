from fpdf import FPDF
import openpyxl
from tkinter import Tk, filedialog
from datetime import datetime

# --- Funções para formatação de dados ---
def formatar_moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"

def formatar_porcentagem(valor):
    try:
        return f"{float(valor) * 100:.2f}%"
    except (ValueError, TypeError):
        return "0%"

def tratar_valor(valor):
    try:
        return int(valor) if valor is not None else 0
    except (ValueError, TypeError):
        return 0

def tratar_porcentagem(valor):
    try:
        return float(valor) if valor is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

# --- Função para seleção de arquivos ---
def selecionar_arquivos():
    Tk().withdraw()
    caminho_planilha = filedialog.askopenfilename(title="Selecione a planilha Excel",
                                                filetypes=[("Arquivos Excel", "*.xlsx;*.xls")])
    caminho_relatorio = filedialog.asksaveasfilename(title="Salvar relatório como",
                                                  defaultextension=".pdf",
                                                  filetypes=[("Arquivos PDF", "*.pdf")])
    return caminho_planilha, caminho_relatorio

# --- Seleção de arquivos ---
caminho_planilha, caminho_relatorio = selecionar_arquivos()

# --- Carregamento e extração de dados da planilha ---
wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
ws = wb["Apesc"]

realizados = {
    "Total": tratar_valor(ws["G13"].value),
    "Acordos": tratar_valor(ws["H13"].value),
    "Sem Ônus": tratar_valor(ws["I13"].value),
    "Condição": tratar_porcentagem(ws["J13"].value),
    "Ticket Médio Geral": tratar_valor(ws["K13"].value),
    "Ticket Médio Pagamento": tratar_valor(ws["L13"].value),
}

metas = {
    "Total": tratar_valor(ws["G18"].value),
    "Acordos": tratar_valor(ws["H18"].value),
    "Sem Ônus": tratar_valor(ws["I18"].value),
    "Condição": tratar_porcentagem(ws["J18"].value),
    "Ticket Médio Geral": tratar_valor(ws["K18"].value),
    "Ticket Médio Pagamento": tratar_valor(ws["L18"].value),
}

# --- Criação do PDF ---
pdf = FPDF()
pdf.set_auto_page_break(auto=False)  # Desabilita a quebra automática de página
pdf.add_page()

# --- Título com data ---
pdf.set_font("Arial", style="B", size=18)
pdf.set_fill_color(50, 50, 150)
pdf.set_text_color(255, 255, 255)
pdf.cell(200, 12, f"Relatório - Apesc {datetime.now().strftime('%d/%m/%Y')}", ln=True, align="C", fill=True)
pdf.ln(10)

# --- Função para adicionar seção com estilo ---
def adicionar_secao(titulo, dados_realizados, dados_metas):
    pdf.set_font("Arial", style="B", size=14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, titulo, ln=True, align="L")
    pdf.set_draw_color(50, 50, 150)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    pdf.set_font("Arial", size=11)

    for chave in dados_realizados.keys():
        realizado = dados_realizados[chave]
        meta = dados_metas[chave]

        # Definindo se atingiu a meta
        if chave in ["Total", "Acordos", "Sem Ônus", "Condição"]:
            diferenca = realizado - meta
            positivo = diferenca >= 0
        else:  # Ticket Médio Geral e Pagamento
            diferenca = meta - realizado
            positivo = diferenca >= 0

        # Estilização do resultado
        pdf.set_font("Arial", style="B", size=11)
        pdf.cell(50, 8, f"{chave}:", border=0)

        pdf.set_font("Arial", size=11)
        if "Ticket" in chave:
            valor_formatado = formatar_moeda(realizado)
        elif "Condição" in chave:
            valor_formatado = formatar_porcentagem(realizado)
        else:
            valor_formatado = str(realizado)

        pdf.cell(40, 8, valor_formatado, border=0)

        pdf.set_font("Arial", style="B", size=11)
        if "Ticket" in chave:
            meta_formatada = formatar_moeda(meta)
        elif "Condição" in chave:
            meta_formatada = formatar_porcentagem(meta)
        else:
            meta_formatada = str(meta)

        pdf.cell(50, 8, f"Meta: {meta_formatada}", border=0)

        # Cor do texto para indicar se atingiu a meta
        pdf.set_text_color(0, 150, 0) if positivo else pdf.set_text_color(200, 0, 0)

        if "Ticket" in chave:
            diferenca_formatada = formatar_moeda(diferenca)
        elif "Condição" in chave:
            diferenca_formatada = formatar_porcentagem(diferenca)
        else:
            diferenca_formatada = str(diferenca)

        pdf.cell(50, 8, f"Variação: {diferenca_formatada}", ln=True)

        # Resetando a cor para preto após cada linha
        pdf.set_text_color(0, 0, 0)

    pdf.ln(8)

# --- Adiciona os dados ao relatório ---
adicionar_secao("Resultados Realizados vs Metas", realizados, metas)

# --- Cálculo da altura do rodapé ---
altura_rodape = 20  # Ajuste conforme necessário

# --- Verificação do espaço disponível na página ---
espaco_restante = pdf.h - pdf.get_y()

# --- Adiciona o rodapé ---
if espaco_restante > altura_rodape:
    pdf.set_y(-altura_rodape)  # Posiciona no final da página
    pdf.set_font("Arial", style="I", size=10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y')} às {datetime.now().strftime('%H:%M')}", align="C", ln=True)
    pdf.cell(0, 10, "Confidencial - Apenas para uso interno", align="C")
else:
    pdf.add_page()  # Cria nova página
    pdf.set_y(-altura_rodape)  # Posiciona no final da página
    pdf.set_font("Arial", style="I", size=10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y')} às {datetime.now().strftime('%H:%M')}", align="C", ln=True)
    pdf.cell(0, 10, "Confidencial - Apenas para uso interno", align="C")


pdf.output(caminho_relatorio)